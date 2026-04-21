# -*- coding: utf-8 -*-
"""查看 xlsx 文件结构"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import load_workbook

FILES = [
    r'e:\Orange\小五云平台\iterations\2026-Q2_结算系统_逾期结算处理优化\testcase\data\smoke_import_2026-01.xlsx',
    r'e:\Orange\小五云平台\iterations\2026-Q2_结算系统_逾期结算处理优化\testcase\data\smoke_import_2026-01_v2.xlsx',
]

for f in FILES:
    print('=' * 80)
    print(f)
    print('=' * 80)
    wb = load_workbook(f, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print('\nSheet: {} (rows={}, cols={})'.format(sheet, ws.max_row, ws.max_column))
        for row in ws.iter_rows(values_only=True):
            print('  {}'.format(row))
    print()
