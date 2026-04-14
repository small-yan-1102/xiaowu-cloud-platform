# -*- coding: utf-8 -*-
"""
AMS 视频下架模块 - 测试数据 Excel 文件生成脚本

根据 ams_video_takedown_data_config.md 中 section 5 的文件清单，
批量生成所有可提前构造的 Excel/CSV 测试文件。

生成目录: testcase/data/excel/
"""

import os
import sys
import random
import string
import csv

# 确保使用项目 Python 环境中的 openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# ============================================================
# 常量定义
# ============================================================

# 输出目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "excel")

# 按作品导入模板列头
WORKS_TEMPLATE_HEADERS = ["作品名称", "CP名称", "海外频道ID"]

# 按视频ID导入模板列头
VIDEO_TEMPLATE_HEADERS = ["视频ID", "作品名称", "CP名称", "海外频道ID"]

# ---------- 真实业务数据（来自 data_config §2） ----------

# HELLO BEAR 分销商的已分配作品（真实数据）
HELLO_BEAR_WORKS = [
    ("版权采买作品001", "严晨芝"),
    ("我叫阮甜甜", "待确认CP"),
    ("麻辣婆媳逗", "待确认CP"),
    ("整个娱乐圈都在等我们离婚", "待确认CP"),
    ("离婚后，我成为顶级神豪", "待确认CP"),
    ("我的合约恋人", "待确认CP"),
    ("无情之家", "待确认CP"),
    ("山之影·花好月圆", "待确认CP"),
    ("植物人老爸是战神", "待确认CP"),
    ("剑影江湖", "待确认CP"),
]

# 核心测试作品
CORE_WORKS = [
    ("亚历山大moto", "境外一号南盾小饼干"),
    ("版权采买作品001", "严晨芝"),
]

# HELLO BEAR 频道（真实数据）
HELLO_BEAR_CHANNELS = [
    "UC_6DWY36uuEWIMHcjPm7UOA",
    "UC_99og720N1KNeSFOABloog",
    "UC_EAmij4Ic67MrbwNe8vNww",
    "UC_Hg6yNn9EHfaY5NLqt6C-w",
    "UC--DohYujPVPdgvYqxZqoFw",
]

# 备用已知视频ID（来自 ams_publish_video）
KNOWN_VIDEO_IDS = [
    "v0vYHYgPC2U",
    "4N-nVkqTgds",
    "NLkJ9ktYLNs",
    "tO6rS5h_3rw",
    "m1pU8xKSEZM",
]


def ensure_output_dir():
    """确保输出目录存在"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    print(f"[INFO] 输出目录: {OUTPUT_DIR}")


def random_video_id(length=11):
    """生成随机的 YouTube 风格视频ID（11位字符）"""
    chars = string.ascii_letters + string.digits + "-_"
    return "".join(random.choice(chars) for _ in range(length))


def random_work_name():
    """生成随机作品名称"""
    prefixes = ["测试作品", "示例作品", "模拟作品", "虚拟作品", "数据作品"]
    return random.choice(prefixes) + str(random.randint(1000, 9999))


def random_cp_name():
    """生成随机 CP 名称"""
    prefixes = ["测试CP", "模拟CP", "虚拟CP", "数据CP"]
    return random.choice(prefixes) + str(random.randint(100, 999))


def random_channel_id():
    """生成随机海外频道ID（YouTube UCxxxx 格式）"""
    chars = string.ascii_letters + string.digits + "-_"
    return "UC" + "".join(random.choice(chars) for _ in range(22))


def save_workbook(wb, filename):
    """保存工作簿到输出目录"""
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    size = os.path.getsize(filepath)
    size_str = f"{size / 1024 / 1024:.2f}MB" if size > 1024 * 1024 else f"{size / 1024:.1f}KB"
    print(f"  [OK] {filename} ({size_str})")
    return filepath


# ============================================================
# Batch 1: 可提前构造的异常校验文件（不依赖业务数据）
# ============================================================

def gen_wrong_template():
    """
    wrong_template.xlsx - VT-021 模板错误校验
    列名/格式不符合导入模板
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 使用错误的列名
    ws.append(["编号", "标题", "描述", "状态"])
    ws.append(["001", "测试数据1", "这是一条错误模板数据", "正常"])
    ws.append(["002", "测试数据2", "这是另一条错误模板数据", "正常"])
    save_workbook(wb, "wrong_template.xlsx")


def gen_oversize_5mb():
    """
    oversize_5mb.xlsx - VT-021 文件大小超限（>5MB）
    使用作品导入模板格式但填充大量数据使文件超过 5MB
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    
    # xlsx 有压缩，需要足够多的随机数据才能超过 5MB
    # 每行约 100 字节压缩后，需要约 60000+ 行
    row_count = 80000
    print(f"  [INFO] 生成 {row_count} 行数据，请稍候...")
    for i in range(row_count):
        ws.append([
            random_work_name() + "_" + "".join(random.choices(string.ascii_letters, k=20)),
            random_cp_name() + "_" + "".join(random.choices(string.ascii_letters, k=20)),
            random_channel_id(),
        ])
    save_workbook(wb, "oversize_5mb.xlsx")


def gen_works_501():
    """
    works_501.xlsx - VT-031 条数边界超限（>500条）
    501 条作品数据，使用作品导入模板格式
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    for i in range(501):
        ws.append([
            f"边界测试作品{i + 1:04d}",
            f"测试CP{(i % 50) + 1:03d}",
            random_channel_id(),
        ])
    save_workbook(wb, "works_501.xlsx")


def gen_invalid_format_csv():
    """
    invalid_format.csv - VT-053 非xlsx格式拦截
    CSV格式文件，内容有效但格式不对
    """
    filepath = os.path.join(OUTPUT_DIR, "invalid_format.csv")
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(WORKS_TEMPLATE_HEADERS)
        writer.writerow(["亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
        writer.writerow(["版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[1]])
    size = os.path.getsize(filepath)
    print(f"  [OK] invalid_format.csv ({size}B)")


def gen_oversize_6mb():
    """
    oversize_6mb.xlsx - VT-053 文件大小超限（>5MB）
    有效xlsx格式但大小超过5MB
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    
    row_count = 100000
    print(f"  [INFO] 生成 {row_count} 行数据，请稍候...")
    for i in range(row_count):
        ws.append([
            random_work_name() + "_" + "".join(random.choices(string.ascii_letters, k=25)),
            random_cp_name() + "_" + "".join(random.choices(string.ascii_letters, k=25)),
            random_channel_id(),
        ])
    save_workbook(wb, "oversize_6mb.xlsx")


def gen_wrong_template_works():
    """
    wrong_template_works.xlsx - VT-053 模板格式错误
    列名为"名称/编号/频道"（非标准模板：作品名称/CP名称/海外频道ID）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 错误的列名
    ws.append(["名称", "编号", "频道"])
    ws.append(["亚历山大moto", "19584", "UC_6DWY36uuEWIMHcjPm7UOA"])
    ws.append(["版权采买作品001", "16425", "UC_99og720N1KNeSFOABloog"])
    save_workbook(wb, "wrong_template_works.xlsx")


def gen_duplicate_works():
    """
    duplicate_works.xlsx - VT-053 文件内作品重复
    包含2行相同作品名称的数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    # 两行完全相同的作品名称
    ws.append(["亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    ws.append(["亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[1]])
    ws.append(["版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[2]])
    save_workbook(wb, "duplicate_works.xlsx")


def gen_videos_1001():
    """
    videos_1001.xlsx - VT-052 条数边界超限（>1000条视频ID）
    1001条视频ID数据，每条为11位字符
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    for i in range(1001):
        ws.append([
            random_video_id(11),
            f"边界视频测试作品{i + 1:04d}",
            f"测试CP{(i % 50) + 1:03d}",
            random_channel_id(),
        ])
    save_workbook(wb, "videos_1001.xlsx")


def gen_vid_wrong_format_csv():
    """
    vid_wrong_format.csv - VT-056 格式校验（CSV非xlsx）
    """
    filepath = os.path.join(OUTPUT_DIR, "vid_wrong_format.csv")
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(VIDEO_TEMPLATE_HEADERS)
        writer.writerow([KNOWN_VIDEO_IDS[0], "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    size = os.path.getsize(filepath)
    print(f"  [OK] vid_wrong_format.csv ({size}B)")


def gen_vid_oversize():
    """
    vid_oversize.xlsx - VT-056 大小校验（>5MB）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    
    row_count = 80000
    print(f"  [INFO] 生成 {row_count} 行数据，请稍候...")
    for i in range(row_count):
        ws.append([
            random_video_id(11),
            random_work_name() + "_" + "".join(random.choices(string.ascii_letters, k=20)),
            random_cp_name() + "_" + "".join(random.choices(string.ascii_letters, k=20)),
            random_channel_id(),
        ])
    save_workbook(wb, "vid_oversize.xlsx")


def gen_vid_wrong_template():
    """
    vid_wrong_template.xlsx - VT-056 模板不匹配（缺少"视频ID"列）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # 缺少"视频ID"列，只有后3列
    ws.append(["作品名称", "CP名称", "海外频道ID"])
    ws.append(["亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    ws.append(["版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "vid_wrong_template.xlsx")


def gen_vid_dup_ids():
    """
    vid_dup_ids.xlsx - VT-056 文件内重复（2行相同视频ID）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    dup_vid = random_video_id(11)
    ws.append([dup_vid, "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    ws.append([dup_vid, "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])  # 重复行
    ws.append([random_video_id(11), "版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "vid_dup_ids.xlsx")


# ============================================================
# Batch 2: 基于模板格式的校验文件（使用无效数据）
# ============================================================

def gen_vid_bad_work():
    """
    vid_bad_work.xlsx - VT-056 作品名称无效
    模板格式正确但作品名称在系统中不存在
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    ws.append([random_video_id(11), "不存在的作品999", "不存在的CP", HELLO_BEAR_CHANNELS[0]])
    ws.append([random_video_id(11), "虚假作品XYZ", "虚假CP", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "vid_bad_work.xlsx")


def gen_vid_bad_cp():
    """
    vid_bad_cp.xlsx - VT-056 CP名称无效
    作品名称有效但CP名称不匹配
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    ws.append([random_video_id(11), "亚历山大moto", "不存在CP", HELLO_BEAR_CHANNELS[0]])
    ws.append([random_video_id(11), "版权采买作品001", "错误CP名称", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "vid_bad_cp.xlsx")


def gen_vid_bad_channel():
    """
    vid_bad_channel.xlsx - VT-056 频道ID不存在
    频道ID格式正确但在系统中不存在
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    ws.append([random_video_id(11), "亚历山大moto", "境外一号南盾小饼干", "000000000"])
    ws.append([random_video_id(11), "版权采买作品001", "严晨芝", "UC_INVALID_CHANNEL_000000"])
    save_workbook(wb, "vid_bad_channel.xlsx")


def gen_vid_no_video():
    """
    vid_no_video.xlsx - VT-056 频道内视频不存在
    频道ID有效但该频道下无此视频
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    # 使用有效频道但随机视频ID（几乎不可能存在）
    ws.append([random_video_id(11), "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    ws.append([random_video_id(11), "版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "vid_no_video.xlsx")


def gen_video_format_test():
    """
    video_format_test.xlsx - VT-045 视频ID 11位边界
    3条数据：1条10字符ID + 1条11字符ID + 1条12字符ID
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    ws.append([random_video_id(10), "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    ws.append([random_video_id(11), "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    ws.append([random_video_id(12), "版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "video_format_test.xlsx")


# ============================================================
# Batch 3: 使用真实业务数据的有效导入文件
# ============================================================

def gen_works_import_3():
    """
    works_import_3.xlsx - VT-022 首次导入（3条合法作品）
    使用 HELLO BEAR 分销商的已分配作品数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    # 使用前3个 HELLO BEAR 作品
    for i, (name, cp) in enumerate(HELLO_BEAR_WORKS[:3]):
        ws.append([name, cp, HELLO_BEAR_CHANNELS[i % len(HELLO_BEAR_CHANNELS)]])
    save_workbook(wb, "works_import_3.xlsx")


def gen_works_import_2():
    """
    works_import_2.xlsx - VT-022 二次导入覆盖（2条合法作品）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    # 使用不同的2个作品（避免与 works_import_3 完全重叠，以验证覆盖效果）
    for i, (name, cp) in enumerate(HELLO_BEAR_WORKS[3:5]):
        ws.append([name, cp, HELLO_BEAR_CHANNELS[i % len(HELLO_BEAR_CHANNELS)]])
    save_workbook(wb, "works_import_2.xlsx")


def gen_partial_fail_5():
    """
    partial_fail_5.xlsx - VT-021 部分失败
    5条数据：3条合法 + 2条作品名不存在
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    # 3条合法数据
    for i, (name, cp) in enumerate(HELLO_BEAR_WORKS[:3]):
        ws.append([name, cp, HELLO_BEAR_CHANNELS[i % len(HELLO_BEAR_CHANNELS)]])
    # 2条不存在的作品
    ws.append(["不存在的测试作品AAA", "虚假CP001", HELLO_BEAR_CHANNELS[0]])
    ws.append(["不存在的测试作品BBB", "虚假CP002", HELLO_BEAR_CHANNELS[1]])
    save_workbook(wb, "partial_fail_5.xlsx")


def gen_works_import_2_for_054():
    """
    works_import_2_for_054.xlsx - VT-054 导入作品 checkbox 禁用验证
    2条有效作品数据
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    for i, (name, cp) in enumerate(HELLO_BEAR_WORKS[5:7]):
        ws.append([name, cp, HELLO_BEAR_CHANNELS[i % len(HELLO_BEAR_CHANNELS)]])
    save_workbook(wb, "works_import_2_for_054.xlsx")


def gen_video_import_valid_3():
    """
    video_import_valid_3.xlsx - VT-006 按视频ID创建E2E
    3条合法视频ID数据，频道需归属 HELLO BEAR

    注意：使用备用视频ID，实际执行时可能需要替换为真实有效的视频ID
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    # 使用已知的视频ID和 HELLO BEAR 作品/频道
    test_data = [
        (KNOWN_VIDEO_IDS[0], "版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[0]),
        (KNOWN_VIDEO_IDS[1], "版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[1]),
        (KNOWN_VIDEO_IDS[2], "版权采买作品001", "严晨芝", HELLO_BEAR_CHANNELS[2]),
    ]
    for row in test_data:
        ws.append(list(row))
    save_workbook(wb, "video_import_valid_3.xlsx")


def gen_video_import_fail():
    """
    video_import_fail.xlsx - VT-024 视频ID校验
    含3种错误：1条非11位ID + 1条重复ID + 1条执行中任务的ID

    注意：第3条"执行中任务的ID"需在 VT-005/006 创建任务后替换
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(VIDEO_TEMPLATE_HEADERS)
    dup_id = random_video_id(11)
    # 1条非11位ID（8位）
    ws.append(["abcd1234", "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    # 1条正常ID（第一次出现，作为重复检测的参照）
    ws.append([dup_id, "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    # 1条重复ID（与上一条相同）
    ws.append([dup_id, "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    # 1条执行中任务的视频ID（占位，执行时需替换）
    ws.append(["PLACEHOLDER1", "亚历山大moto", "境外一号南盾小饼干", HELLO_BEAR_CHANNELS[0]])
    save_workbook(wb, "video_import_fail.xlsx")


def gen_works_500_valid():
    """
    works_500_valid.xlsx - VT-067 大批量500条作品性能验证
    500条有效作品数据

    注意：理想情况应从 ams_composition (874条) 中导出真实数据。
    当前使用 HELLO BEAR 已分配作品 + 模拟数据填充至500条。
    执行前建议替换为系统中的真实作品数据。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(WORKS_TEMPLATE_HEADERS)
    
    # 先填入已知真实作品
    for i, (name, cp) in enumerate(HELLO_BEAR_WORKS):
        ws.append([name, cp, HELLO_BEAR_CHANNELS[i % len(HELLO_BEAR_CHANNELS)]])
    
    # 剩余用模拟数据填充（执行前建议替换为真实数据）
    remaining = 500 - len(HELLO_BEAR_WORKS)
    for i in range(remaining):
        ws.append([
            f"批量测试作品{i + 1:04d}",
            f"批量CP{(i % 100) + 1:03d}",
            HELLO_BEAR_CHANNELS[i % len(HELLO_BEAR_CHANNELS)],
        ])
    save_workbook(wb, "works_500_valid.xlsx")


# ============================================================
# 主函数
# ============================================================

def main():
    """主入口：按批次生成所有测试文件"""
    
    print("=" * 60)
    print("AMS 视频下架模块 - 测试数据 Excel 文件生成")
    print("=" * 60)
    
    ensure_output_dir()
    
    # ---------- Batch 1: 异常校验文件（不依赖业务数据） ----------
    print("\n--- Batch 1: 异常校验文件（不依赖业务数据） ---")
    
    print("[1/12] wrong_template.xlsx (VT-021)")
    gen_wrong_template()
    
    print("[2/12] oversize_5mb.xlsx (VT-021)")
    gen_oversize_5mb()
    
    print("[3/12] works_501.xlsx (VT-031)")
    gen_works_501()
    
    print("[4/12] invalid_format.csv (VT-053)")
    gen_invalid_format_csv()
    
    print("[5/12] oversize_6mb.xlsx (VT-053)")
    gen_oversize_6mb()
    
    print("[6/12] wrong_template_works.xlsx (VT-053)")
    gen_wrong_template_works()
    
    print("[7/12] duplicate_works.xlsx (VT-053)")
    gen_duplicate_works()
    
    print("[8/12] videos_1001.xlsx (VT-052)")
    gen_videos_1001()
    
    print("[9/12] vid_wrong_format.csv (VT-056)")
    gen_vid_wrong_format_csv()
    
    print("[10/12] vid_oversize.xlsx (VT-056)")
    gen_vid_oversize()
    
    print("[11/12] vid_wrong_template.xlsx (VT-056)")
    gen_vid_wrong_template()
    
    print("[12/12] vid_dup_ids.xlsx (VT-056)")
    gen_vid_dup_ids()
    
    # ---------- Batch 2: 基于模板格式的校验文件 ----------
    print("\n--- Batch 2: 基于模板格式的校验文件（使用无效数据） ---")
    
    print("[1/5] vid_bad_work.xlsx (VT-056)")
    gen_vid_bad_work()
    
    print("[2/5] vid_bad_cp.xlsx (VT-056)")
    gen_vid_bad_cp()
    
    print("[3/5] vid_bad_channel.xlsx (VT-056)")
    gen_vid_bad_channel()
    
    print("[4/5] vid_no_video.xlsx (VT-056)")
    gen_vid_no_video()
    
    print("[5/5] video_format_test.xlsx (VT-045)")
    gen_video_format_test()
    
    # ---------- Batch 3: 使用真实业务数据的导入文件 ----------
    print("\n--- Batch 3: 使用真实业务数据的导入文件 ---")
    
    print("[1/7] works_import_3.xlsx (VT-022)")
    gen_works_import_3()
    
    print("[2/7] works_import_2.xlsx (VT-022)")
    gen_works_import_2()
    
    print("[3/7] partial_fail_5.xlsx (VT-021)")
    gen_partial_fail_5()
    
    print("[4/7] works_import_2_for_054.xlsx (VT-054)")
    gen_works_import_2_for_054()
    
    print("[5/7] video_import_valid_3.xlsx (VT-006)")
    gen_video_import_valid_3()
    
    print("[6/7] video_import_fail.xlsx (VT-024)")
    gen_video_import_fail()
    
    print("[7/7] works_500_valid.xlsx (VT-067)")
    gen_works_500_valid()
    
    # ---------- 汇总 ----------
    print("\n" + "=" * 60)
    total_files = len([f for f in os.listdir(OUTPUT_DIR) if f.endswith(('.xlsx', '.csv'))])
    print(f"生成完成！共 {total_files} 个文件，存放于: {OUTPUT_DIR}")
    
    # 列出不包含在本次生成中的文件
    print("\n--- 需要运行时构造的文件（本次未生成） ---")
    print("  - vid_valid_alloc.xlsx (VT-057) 需查询分配关系后构造")
    print("  - vid_invalid_alloc.xlsx (VT-057) 需查询分配关系后构造")
    print("  - vid_valid_ops.xlsx (VT-057) 需查询分配关系后构造")
    print("  - vid_invalid_ops.xlsx (VT-057) 需查询分配关系后构造")
    
    print("\n--- 执行前需确认/替换的文件 ---")
    print("  - video_import_valid_3.xlsx: 视频ID需确认在HELLO BEAR频道下有效")
    print("  - video_import_fail.xlsx: PLACEHOLDER1 需替换为执行中任务的视频ID")
    print("  - works_500_valid.xlsx: 建议用真实作品数据替换模拟数据")
    print("  - oversize_5mb.xlsx / oversize_6mb.xlsx / vid_oversize.xlsx: 确认大小>5MB")


if __name__ == "__main__":
    main()
